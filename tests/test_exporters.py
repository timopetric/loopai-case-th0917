"""`exporters.to_csv` unit tests (issue 10).

The CSV is read back with Python's own `csv` module — a plain consumer, not
the raw string — because the point of "pure data" is that a standard reader
needs no special handling. Content is checked against a `ReportTable` built
by hand (not through `engine.execute`) so these tests describe the exporter's
contract in isolation from aggregation correctness, which `test_engine.py`
already covers.
"""

import csv
import io
import json

import openpyxl
import pytest

from app.assumptions import AssumptionNote, build_assumptions
from app.engine import execute
from app.exporters import to_csv, to_xlsx
from app.models import ColumnMeta, Metric, ReportRow, ReportSpec, ReportTable
from app.upstream import _DEV_FIXTURE_PATH, CoverageWindow, _normalise_dataset

COVERAGE = CoverageWindow(from_date="2026-07-10", to_date="2026-07-23")


def _spec(**overrides) -> ReportSpec:
    body = dict(
        metrics=["resolved"],
        date_from="2026-07-10",
        date_to="2026-07-23",
        granularity="day",
        group_by="none",
    )
    body.update(overrides)
    return ReportSpec(**body)


def _read_csv(text: str) -> list[list[str]]:
    return list(csv.reader(io.StringIO(text)))


def test_csv_parses_with_a_standard_reader_and_has_a_header_plus_one_row_per_report_row() -> None:
    table = ReportTable(
        columns=[ColumnMeta(key="resolved", label="Resolved", kind="counter", unit="count")],
        rows=[
            ReportRow(
                bucket="2026-07-10", group_key=None, group_label=None, values={"resolved": 5}
            ),
            ReportRow(
                bucket="2026-07-11", group_key=None, group_label=None, values={"resolved": 7}
            ),
        ],
        totals={"resolved": 12},
    )

    text = to_csv(_spec(), table)
    rows = _read_csv(text)

    # header, two data rows, one totals row — no preamble above the header.
    assert len(rows) == 4
    assert rows[0][0] == "Day"
    assert rows[1] == ["2026-07-10", "5"]
    assert rows[2] == ["2026-07-11", "7"]
    assert rows[3] == ["Total", "12"]


def test_a_duration_metric_column_header_names_its_unit() -> None:
    table = ReportTable(
        columns=[
            ColumnMeta(key="resolve_time", label="Resolve time", kind="duration", unit="hours")
        ],
        rows=[
            ReportRow(
                bucket="2026-07-10", group_key=None, group_label=None, values={"resolve_time": 1.5}
            )
        ],
        totals={"resolve_time": 1.5},
    )

    rows = _read_csv(to_csv(_spec(metrics=["resolve_time"]), table))

    assert rows[0] == ["Day", "Resolve time (h)"]


def test_a_withheld_total_renders_as_an_empty_field_not_a_dash_and_not_zero() -> None:
    """The non-additive rule (architecture.md "Table semantics"): `actioned_emails`
    totalled across Actors is withheld (`None`). In CSV — the machine-readable
    format — that renders as an empty field, not the on-screen dash: an
    em dash in a numeric column forces a script's parser to treat the whole
    column as text (user story 33, architecture.md §12). Never a lying `0`
    either way."""
    table = ReportTable(
        columns=[
            ColumnMeta(key="actioned_emails", label="Actioned emails", kind="counter", unit="count")
        ],
        rows=[
            ReportRow(
                bucket="2026-07-10",
                group_key="a1",
                group_label="Alice",
                values={"actioned_emails": 3},
            )
        ],
        totals={"actioned_emails": None},
    )

    rows = _read_csv(to_csv(_spec(metrics=["actioned_emails"], group_by="agent"), table))

    assert rows[0] == ["Day", "Actor", "Actioned emails"]
    assert rows[1] == ["2026-07-10", "Alice", "3"]
    totals_row = rows[-1]
    assert totals_row[-1] == ""
    assert totals_row[-1] != "—"
    assert totals_row[-1] != "0"


def test_a_zero_count_duration_average_in_an_ordinary_data_row_is_an_empty_field() -> None:
    """The withheld sentinel isn't only a totals-row concern: a zero-ticket
    Actor's Duration Metric average is `None` on an ordinary data row too
    (`engine._display_value`), and that is exactly where 47 of 108 rows
    carried a stray dash before this fix — the totals row alone never
    exercised that path."""
    table = ReportTable(
        columns=[
            ColumnMeta(key="resolved", label="Resolved", kind="counter", unit="count"),
            ColumnMeta(key="resolve_time", label="Resolve time", kind="duration", unit="hours"),
        ],
        rows=[
            ReportRow(
                bucket="total",
                group_key="a1",
                group_label="Idle Actor",
                values={"resolved": 0, "resolve_time": None},
            ),
            ReportRow(
                bucket="total",
                group_key="a2",
                group_label="Busy Actor",
                values={"resolved": 4, "resolve_time": 2.5},
            ),
        ],
        totals={"resolved": 4, "resolve_time": 2.5},
    )

    rows = _read_csv(
        to_csv(_spec(metrics=["resolved", "resolve_time"], group_by="agent"), table)
    )

    idle_row = next(r for r in rows if r[1] == "Idle Actor")
    busy_row = next(r for r in rows if r[1] == "Busy Actor")
    assert idle_row[-1] == ""
    assert busy_row[-1] == "2.5"


def test_every_value_in_a_numeric_column_is_empty_or_parses_as_a_float() -> None:
    """The §12 property directly, without adding pandas as a dependency:
    `float()` over every non-empty cell in a metric column must not raise —
    the same guarantee `pandas.read_csv(...)[col].astype(float)` depends on.
    A mix of real numbers and a withheld zero-count average in the same
    column is exactly the shape that broke before this fix."""
    table = ReportTable(
        columns=[
            ColumnMeta(key="resolve_time", label="Resolve time", kind="duration", unit="hours")
        ],
        rows=[
            ReportRow(
                bucket="total", group_key="a1", group_label="Idle", values={"resolve_time": None}
            ),
            ReportRow(
                bucket="total", group_key="a2", group_label="Busy", values={"resolve_time": 3.25}
            ),
        ],
        totals={"resolve_time": None},
    )

    rows = _read_csv(to_csv(_spec(metrics=["resolve_time"], group_by="agent"), table))

    header, *data_and_totals = rows
    column_index = header.index("Resolve time (h)")
    cells = [r[column_index] for r in data_and_totals]

    assert cells == ["", "3.25", ""]
    for cell in cells:
        if cell != "":
            float(cell)  # must not raise


def test_a_group_label_containing_a_comma_and_a_quote_round_trips_through_csv() -> None:
    """CSV correctness failures are usually escaping failures — a group label
    like a real Actor name can contain a comma or an embedded quote, and the
    stdlib `csv` reader must hand back the exact original string."""
    tricky_name = 'Smith, Jane "JJ"'
    table = ReportTable(
        columns=[ColumnMeta(key="resolved", label="Resolved", kind="counter", unit="count")],
        rows=[
            ReportRow(
                bucket="2026-07-10", group_key="a1", group_label=tricky_name, values={"resolved": 2}
            )
        ],
        totals={"resolved": 2},
    )

    rows = _read_csv(to_csv(_spec(group_by="agent"), table))

    assert rows[1][1] == tricky_name


def test_column_order_follows_report_table_columns_not_metrics_order() -> None:
    """`engine._ordered_metrics` already applies `columns_order` to build
    `ReportTable.columns` (issue 07) — the exporter just has to read that
    list in the order given, not re-derive an order of its own."""
    table = ReportTable(
        columns=[
            ColumnMeta(key="new_tickets", label="New tickets", kind="counter", unit="count"),
            ColumnMeta(key="resolved", label="Resolved", kind="counter", unit="count"),
        ],
        rows=[
            ReportRow(
                bucket="2026-07-10",
                group_key=None,
                group_label=None,
                values={"resolved": 5, "new_tickets": 9},
            )
        ],
        totals={"resolved": 5, "new_tickets": 9},
    )

    rows = _read_csv(to_csv(_spec(metrics=["resolved", "new_tickets"]), table))

    assert rows[0] == ["Day", "New tickets", "Resolved"]
    assert rows[1] == ["2026-07-10", "9", "5"]


def test_csv_content_matches_the_real_report_table_from_the_committed_fixture() -> None:
    """The exporter derives from the same `ReportTable` the preview renders
    (issue 10's core guarantee) — proved here with the real fixture dataset,
    not a hand-built table, going through `engine.execute` exactly as the
    `/report` route does."""
    dataset = _normalise_dataset(
        json.loads(_DEV_FIXTURE_PATH.read_text())["response_json"],
        CoverageWindow(from_date="2026-07-10", to_date="2026-07-23"),
    )
    spec = ReportSpec(
        metrics=[Metric.RESOLVED],
        date_from="2026-07-10",
        date_to="2026-07-23",
        granularity="total",
        group_by="none",
    )

    table = execute(spec, dataset)
    rows = _read_csv(to_csv(spec, table))

    assert rows[0] == ["Day", "Resolved"]
    assert rows[1] == ["total", str(int(table.rows[0].values["resolved"]))]
    assert rows[1][1] == "16372"
    assert rows[2] == ["Total", "16372"]


# --- to_xlsx (issue 11) -----------------------------------------------------


def _workbook(spec: ReportSpec, table: ReportTable, coverage: CoverageWindow = COVERAGE):
    return openpyxl.load_workbook(io.BytesIO(to_xlsx(spec, table, coverage)))


def test_workbook_has_a_data_sheet_and_a_report_info_sheet() -> None:
    table = ReportTable(
        columns=[ColumnMeta(key="resolved", label="Resolved", kind="counter", unit="count")],
        rows=[
            ReportRow(
                bucket="2026-07-10", group_key=None, group_label=None, values={"resolved": 5}
            )
        ],
        totals={"resolved": 5},
    )

    wb = _workbook(_spec(), table)

    assert wb.sheetnames == ["Data", "Report info"]


def test_data_sheet_matches_the_csv_except_the_withheld_encoding() -> None:
    """The data sheet is laid out identically to the CSV — same header, same
    rows, same totals row — with exactly one deliberate difference: a
    withheld cell is an em dash here, an empty field in the CSV (module
    docstring's documented split, not a discrepancy)."""
    table = ReportTable(
        columns=[
            ColumnMeta(key="actioned_emails", label="Actioned emails", kind="counter", unit="count")
        ],
        rows=[
            ReportRow(
                bucket="2026-07-10",
                group_key="a1",
                group_label="Alice",
                values={"actioned_emails": 3},
            )
        ],
        totals={"actioned_emails": None},
    )
    spec = _spec(metrics=["actioned_emails"], group_by="agent")

    csv_rows = _read_csv(to_csv(spec, table))
    wb = _workbook(spec, table)
    data_sheet = wb["Data"]
    xlsx_rows = [[cell.value for cell in row] for row in data_sheet.iter_rows()]

    assert xlsx_rows[0] == csv_rows[0]
    assert xlsx_rows[1] == ["2026-07-10", "Alice", 3]
    assert csv_rows[1] == ["2026-07-10", "Alice", "3"]
    # The withheld total: CSV empty field vs XLSX em dash.
    assert csv_rows[-1][-1] == ""
    assert xlsx_rows[-1][-1] == "—"


def test_duration_values_are_written_as_numeric_cells_not_numeric_looking_strings() -> None:
    """The load-bearing assertion (issue 11): a duration written as text
    looks identical on screen but silently breaks `=AVERAGE()`. Assert the
    cell's Python type read back by openpyxl, not just its printed value."""
    table = ReportTable(
        columns=[
            ColumnMeta(key="resolve_time", label="Resolve time", kind="duration", unit="hours")
        ],
        rows=[
            ReportRow(
                bucket="2026-07-10",
                group_key=None,
                group_label=None,
                values={"resolve_time": 1.5},
            )
        ],
        totals={"resolve_time": 1.5},
    )
    spec = _spec(metrics=["resolve_time"])

    wb = _workbook(spec, table)
    data_sheet = wb["Data"]

    header_row = next(data_sheet.iter_rows(min_row=1, max_row=1))
    assert header_row[-1].value == "Resolve time (h)"

    data_row = next(data_sheet.iter_rows(min_row=2, max_row=2))
    totals_row = next(data_sheet.iter_rows(min_row=3, max_row=3))
    assert isinstance(data_row[-1].value, int | float)
    assert data_row[-1].value == 1.5
    assert isinstance(totals_row[-1].value, int | float)


def test_report_info_sheet_carries_definition_coverage_units_note_and_warnings() -> None:
    """A spec that groups `actioned_emails` by Actor raises the non-additive
    Warning (`engine._totals`) — used here as a real Warning that must
    genuinely reach the sheet, not a hand-built `ReportTable.warnings` list."""
    dataset = _normalise_dataset(
        json.loads(_DEV_FIXTURE_PATH.read_text())["response_json"], COVERAGE
    )
    spec = ReportSpec(
        metrics=[Metric.ACTIONED_EMAILS],
        date_from="2026-07-10",
        date_to="2026-07-23",
        granularity="total",
        group_by="agent",
    )
    table = execute(spec, dataset)
    assert table.warnings  # sanity: this spec really does raise a Warning

    wb = _workbook(spec, table, COVERAGE)
    info_sheet = wb["Report info"]
    text = "\n".join(
        str(cell.value) for row in info_sheet.iter_rows() for cell in row if cell.value is not None
    )

    # Report definition.
    assert "Actioned emails" in text
    assert "2026-07-10" in text and "2026-07-23" in text
    assert "agent" in spec.group_by  # sanity on the fixture spec itself
    assert "Actor" in text  # the human label for group_by == "agent"

    # Coverage Window.
    assert COVERAGE.from_date in text
    assert COVERAGE.to_date in text

    # Units note — the same text `build_assumptions` produces, not a paraphrase.
    units_note = next(n for n in build_assumptions(COVERAGE) if n.id == "units_hours")
    assert units_note.title in text
    assert units_note.body in text

    # The real Warning raised for this spec.
    for warning in table.warnings:
        assert warning in text


def test_report_info_sheet_states_no_warnings_when_none_were_raised() -> None:
    table = ReportTable(
        columns=[ColumnMeta(key="resolved", label="Resolved", kind="counter", unit="count")],
        rows=[
            ReportRow(
                bucket="2026-07-10", group_key=None, group_label=None, values={"resolved": 5}
            )
        ],
        totals={"resolved": 5},
        warnings=[],
    )

    wb = _workbook(_spec(), table)
    info_sheet = wb["Report info"]
    text = "\n".join(
        str(cell.value) for row in info_sheet.iter_rows() for cell in row if cell.value is not None
    )

    assert "None" in text


def test_report_info_sheet_entity_filter_row_carries_the_real_value_when_set() -> None:
    table = ReportTable(
        columns=[ColumnMeta(key="resolved", label="Resolved", kind="counter", unit="count")],
        rows=[
            ReportRow(
                bucket="2026-07-10", group_key=None, group_label=None, values={"resolved": 5}
            )
        ],
        totals={"resolved": 5},
        warnings=[],
    )

    wb = _workbook(_spec(group_by="agent", entity_filter="Alice"), table)
    info_sheet = wb["Report info"]
    rows = [[cell.value for cell in row] for row in info_sheet.iter_rows()]

    grouped_by_index = next(i for i, row in enumerate(rows) if row[0] == "Grouped by")
    assert rows[grouped_by_index + 1] == ["Entity filter", "Alice"]


def test_report_info_sheet_entity_filter_row_reads_none_when_unset() -> None:
    table = ReportTable(
        columns=[ColumnMeta(key="resolved", label="Resolved", kind="counter", unit="count")],
        rows=[
            ReportRow(
                bucket="2026-07-10", group_key=None, group_label=None, values={"resolved": 5}
            )
        ],
        totals={"resolved": 5},
        warnings=[],
    )

    wb = _workbook(_spec(), table)
    info_sheet = wb["Report info"]
    rows = [[cell.value for cell in row] for row in info_sheet.iter_rows()]

    grouped_by_index = next(i for i, row in enumerate(rows) if row[0] == "Grouped by")
    assert rows[grouped_by_index + 1] == ["Entity filter", "None"]


def test_csv_ignores_entity_filter_entirely() -> None:
    """A CSV export must not leak `entity_filter` anywhere in the file — no
    preamble line, no extra column, no header change. The only permitted
    difference from an unfiltered export is which rows are present, and this
    test holds the `ReportTable` fixed to isolate that: same table, only the
    spec's `entity_filter` differs, so the CSV text must be byte-identical."""
    table = ReportTable(
        columns=[ColumnMeta(key="resolved", label="Resolved", kind="counter", unit="count")],
        rows=[
            ReportRow(
                bucket="2026-07-10", group_key="a1", group_label="Alice", values={"resolved": 5}
            )
        ],
        totals={"resolved": 5},
        warnings=[],
    )

    unfiltered = to_csv(_spec(group_by="agent"), table)
    filtered = to_csv(_spec(group_by="agent", entity_filter="Alice"), table)

    assert unfiltered == filtered


def test_report_info_sheet_defers_the_units_note_to_the_shared_assumptions_source(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """The decisive drift-proofing test, mirrored from
    `test_assumptions_route_has_no_hardcoded_second_copy`: patch the shared
    source to return a sentinel `units_hours` note, and require the
    workbook to carry that sentinel text verbatim. A test that only checked
    for the expected sentences would pass just as well against a copy-pasted
    duplicate — this one fails the moment the exporter stops calling
    `build_assumptions`."""
    sentinel = AssumptionNote(
        id="units_hours", title="Sentinel units title", body="Sentinel units body"
    )

    import app.exporters as exporters_module

    monkeypatch.setattr(exporters_module, "build_assumptions", lambda coverage: [sentinel])

    table = ReportTable(
        columns=[ColumnMeta(key="resolved", label="Resolved", kind="counter", unit="count")],
        rows=[
            ReportRow(
                bucket="2026-07-10", group_key=None, group_label=None, values={"resolved": 5}
            )
        ],
        totals={"resolved": 5},
    )

    wb = _workbook(_spec(), table)
    info_sheet = wb["Report info"]
    text = "\n".join(
        str(cell.value) for row in info_sheet.iter_rows() for cell in row if cell.value is not None
    )

    assert sentinel.title in text
    assert sentinel.body in text
