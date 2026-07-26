"""API-level test for the CSV export route (issue 10).

`upstream` is faked via the same FastAPI dependency override pattern as
`test_api_report.py` — no network. These tests cover the seam pure unit
tests on `exporters.to_csv` cannot: that the route is wired up, auth-gated,
returns a spreadsheet content type, and the body a real HTTP response
carries actually parses with a standard CSV reader.
"""

import csv
import io
import json

import openpyxl
import pytest
from fastapi.testclient import TestClient

from app.api.v1.routers.meta import get_upstream_client
from app.config import get_settings
from app.main import create_app
from app.upstream import _DEV_FIXTURE_PATH, CoverageWindow, _normalise_dataset

FIXTURE_PATH = _DEV_FIXTURE_PATH


class _FixtureBackedClient:
    def __init__(self) -> None:
        raw = json.loads(FIXTURE_PATH.read_text())["response_json"]
        self._dataset = _normalise_dataset(
            raw, CoverageWindow(from_date="2026-07-10", to_date="2026-07-23")
        )

    async def get_dataset(self):
        return self._dataset

    async def get_coverage_window(self):
        return self._dataset.coverage


@pytest.fixture
def app():
    application = create_app()
    application.dependency_overrides[get_upstream_client] = lambda: _FixtureBackedClient()
    return application


@pytest.fixture
def client(app) -> TestClient:
    return TestClient(app)


def _auth_headers() -> dict[str, str]:
    return {"X-API-Key": get_settings().app_api_key}


def _spec(**overrides) -> dict:
    body = dict(
        metrics=["resolved"],
        date_from="2026-07-10",
        date_to="2026-07-23",
        granularity="total",
        group_by="none",
    )
    body.update(overrides)
    return body


def test_export_csv_route_returns_a_spreadsheet_content_type_with_a_parsing_body(
    client: TestClient,
) -> None:
    response = client.post("/api/v1/export/csv", json=_spec(), headers=_auth_headers())

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert "attachment" in response.headers["content-disposition"]
    assert response.headers["content-disposition"].endswith('.csv"')

    rows = list(csv.reader(io.StringIO(response.text)))
    assert rows[0] == ["Day", "Resolved"]
    assert rows[1] == ["total", "16372"]
    assert rows[2] == ["Total", "16372"]


def test_export_csv_route_requires_auth(client: TestClient) -> None:
    response = client.post("/api/v1/export/csv", json=_spec())

    assert response.status_code == 401


def test_export_csv_route_rejects_an_invalid_spec(client: TestClient) -> None:
    response = client.post(
        "/api/v1/export/csv",
        json=_spec(date_from="2026-07-23", date_to="2026-07-10"),
        headers=_auth_headers(),
    )

    assert response.status_code == 422


def test_export_xlsx_route_returns_a_spreadsheet_content_type_with_both_sheets(
    client: TestClient,
) -> None:
    response = client.post("/api/v1/export/xlsx", json=_spec(), headers=_auth_headers())

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in response.headers["content-disposition"]
    assert response.headers["content-disposition"].endswith('.xlsx"')

    wb = openpyxl.load_workbook(io.BytesIO(response.content))
    assert wb.sheetnames == ["Data", "Report info"]

    data_rows = list(wb["Data"].iter_rows(values_only=True))
    assert data_rows[0] == ("Day", "Resolved")
    assert data_rows[1] == ("total", 16372)
    assert data_rows[2] == ("Total", 16372)


def test_export_xlsx_route_requires_auth(client: TestClient) -> None:
    response = client.post("/api/v1/export/xlsx", json=_spec())

    assert response.status_code == 401


def test_export_xlsx_route_rejects_an_invalid_spec(client: TestClient) -> None:
    response = client.post(
        "/api/v1/export/xlsx",
        json=_spec(date_from="2026-07-23", date_to="2026-07-10"),
        headers=_auth_headers(),
    )

    assert response.status_code == 422
